"""
Creates a daily excel report from ezee span and demand point layers.
"""
import arcgis
import openpyxl as pyxl
import pandas as pd
import datetime
from datetime import timezone
import os

def datetime_last_week():
    # Get the current time
    current_time = datetime.datetime.now(timezone.utc)

    # Find the start of the current week
    start_of_week = current_time - datetime.timedelta(days=current_time.weekday())

    # Subtract 7 days to get the start of last week
    start_of_last_week = start_of_week - datetime.timedelta(days=7)

    return start_of_last_week

def datetime_this_week():
    # Get the current time
    current_time = datetime.datetime.now(timezone.utc)

    # Find the start of the current week
    start_of_week = current_time - datetime.timedelta(days=current_time.weekday())

    return start_of_week

def datetime_1_day_ago():
    # Get the current time
    current_time = datetime.datetime.now(timezone.utc)

    # Subtract 15 minutes from the current time
    time_1_day_ago = current_time - datetime.timedelta(days=1)

    return time_1_day_ago

def total_length(x):
     return x['CALCULATED_LENGTH'].sum()

def sum_conduit_yesterday(x):
        yesterday_date = datetime_1_day_ago()
        return x[x['Conduit_Placed_Date'].dt.date == yesterday_date]['CALCULATED_LENGTH'].sum()

def main():
    gis = arcgis.gis.GIS("https://ezeefiber.maps.arcgis.com/home", 'cbrown_lightsett', 'Redcar=1')
    span_id = 'ad7e5f94dd5d46e2a7cab5e5a2c6966a'
    span_item = gis.content.get(span_id)
    span_lyr = span_item.layers[0]
    
    print('querying data')
    span_df = span_lyr.query(as_df=True)
    
    con_placed_yesterday_df = span_lyr.query(
        where=f"Conduit_Placed_Date >= TIMESTAMP '{datetime_1_day_ago()}'",
        as_df=True)
    con_placed_this_week_df = span_lyr.query(
        where=f"Conduit_Placed_Date >= TIMESTAMP '{datetime_this_week()}'",
        as_df=True)
    con_placed_last_week_df = span_lyr.query(
        where=f"Conduit_Placed_Date >= TIMESTAMP '{datetime_last_week()}' AND Conduit_Placed_Date <= TIMESTAMP '{datetime_this_week()}'",
        as_df=True)
    con_placed_df = span_lyr.query(
        where=f"Conduit_Placed_Date IS NOT NULL",
        as_df=True)
    
    fiber_placed_yesterday_df = span_lyr.query(
        where=f"Fiber_Placed_Date >= TIMESTAMP '{datetime_1_day_ago()}'",
        as_df=True)
    fiber_placed_this_week_df = span_lyr.query(
        where=f"Fiber_Placed_Date >= TIMESTAMP '{datetime_this_week()}'",
        as_df=True)
    fiber_placed_last_week_df = span_lyr.query(
        where=f"Fiber_Placed_Date >= TIMESTAMP '{datetime_last_week()}' AND Fiber_Placed_Date <= TIMESTAMP '{datetime_this_week()}'",
        as_df=True)
    fiber_placed_df = span_lyr.query(
        where=f"Fiber_Placed_Date IS NOT NULL",
        as_df=True)
    
    """
    # Demand Points
    demand_id = '9b0f2195521548f8b25074f9bd9d9e58'
    demand_item = gis.content.get(demand_id)
    demand_lyr = demand_item.layers[0]
    demand_df = demand_lyr.query(as_df=True)
    
    # Total Demand
    total_demand_df = demand_df.groupby(by=['PROJECT_NAME', 'Work_Order_ID']).size().reset_index(name='HHP')
    total_demand_df.rename(columns={'Work_Order_ID': 'WORK_ORDER_ID'}, inplace=True)
    """

    # Permits
    permit_id = '77217fd167944a05806807052b0a634b'
    permit_item = gis.content.get(permit_id)
    permit_lyr = permit_item.layers[0]
    permit_df = permit_lyr.query(as_df=True)

    permit_df['SUBMITTED_DATE'] = pd.to_datetime(permit_df['SUBMITTED_DATE'])
    permit_df['APPROVED_DATE'] = pd.to_datetime(permit_df['APPROVED_DATE'])

    permit_df['Cycle_Time'] = permit_df['APPROVED_DATE'] - permit_df['SUBMITTED_DATE']

    print('grouping dataframes')
    # total permits
    total_permit_df = permit_df.groupby(by=['APPROVING_AUTHORITY'])['Cycle_Time'].mean().reset_index()
    total_permit_df.rename(columns={'Cycle_Time': 'Avg. Cycle Time'}, inplace=True)

    # activated permits
    filtered_permits = permit_df[permit_df['STATUS'] == 'Activated']
    activated_permits = filtered_permits[['PROJECT_NAME', 'WORK_ORDER_ID', 'STATUS']].copy()

    # Total Length
    total_length_df = span_df.groupby(by=['PROJECT_NAME', 'WORK_ORDER_ID'])['CALCULATED_LENGTH'].sum().reset_index()
    total_length_df.rename(columns={'CALCULATED_LENGTH': 'Total_Length'}, inplace=True)
    
    # Conduit Yesterday
    con_yesterday_length_df = con_placed_yesterday_df.groupby(by=['PROJECT_NAME', 'WORK_ORDER_ID'])['CALCULATED_LENGTH'].sum().reset_index()
    con_yesterday_length_df.rename(columns={'CALCULATED_LENGTH': 'Conduit_Placed_Yesterday'}, inplace=True)

    # Conduit This Week
    con_this_week_length_df = con_placed_this_week_df.groupby(by=['PROJECT_NAME', 'WORK_ORDER_ID'])['CALCULATED_LENGTH'].sum().reset_index()
    con_this_week_length_df.rename(columns={'CALCULATED_LENGTH': 'Conduit_Placed_This_Week'}, inplace=True)

    # Conduit Last Week
    con_last_week_length_df = con_placed_last_week_df.groupby(by=['PROJECT_NAME', 'WORK_ORDER_ID'])['CALCULATED_LENGTH'].sum().reset_index()
    con_last_week_length_df.rename(columns={'CALCULATED_LENGTH': 'Conduit_Placed_Last_Week'}, inplace=True)

    # Conduit Placed
    con_placed_length_df = con_placed_df.groupby(by=['PROJECT_NAME', 'WORK_ORDER_ID'])['CALCULATED_LENGTH'].sum().reset_index()
    con_placed_length_df.rename(columns={'CALCULATED_LENGTH': 'Conduit_Placed'}, inplace=True)

    # Fiber Yesterday
    fiber_yesterday_length_df = fiber_placed_yesterday_df.groupby(by=['PROJECT_NAME', 'WORK_ORDER_ID'])['CALCULATED_LENGTH'].sum().reset_index()
    fiber_yesterday_length_df.rename(columns={'CALCULATED_LENGTH': 'Fiber_Placed_Yesterday'}, inplace=True)

    # Fiber This Week
    fiber_this_week_length_df = fiber_placed_this_week_df.groupby(by=['PROJECT_NAME', 'WORK_ORDER_ID'])['CALCULATED_LENGTH'].sum().reset_index()
    fiber_this_week_length_df.rename(columns={'CALCULATED_LENGTH': 'Fiber_Placed_This_Week'}, inplace=True)

    # Fiber Last Week
    fiber_last_week_length_df = fiber_placed_last_week_df.groupby(by=['PROJECT_NAME', 'WORK_ORDER_ID'])['CALCULATED_LENGTH'].sum().reset_index()
    fiber_last_week_length_df.rename(columns={'CALCULATED_LENGTH': 'Fiber_Placed_Last_Week'}, inplace=True)

    # Fiber Placed
    fiber_placed_length_df = fiber_placed_df.groupby(by=['PROJECT_NAME', 'WORK_ORDER_ID'])['CALCULATED_LENGTH'].sum().reset_index()
    fiber_placed_length_df.rename(columns={'CALCULATED_LENGTH': 'Fiber_Placed'}, inplace=True)

    print('merging')
    merged_df = pd.merge(
        total_length_df, 
        con_yesterday_length_df, 
        on=['PROJECT_NAME', 'WORK_ORDER_ID'], how='outer')
    
    merged_df = pd.merge(
        merged_df, 
        con_this_week_length_df, 
        on=['PROJECT_NAME', 'WORK_ORDER_ID'], how='outer')
    
    merged_df = pd.merge(
        merged_df, 
        con_last_week_length_df, 
        on=['PROJECT_NAME', 'WORK_ORDER_ID'], how='outer')
    
    merged_df = pd.merge(
        merged_df, 
        con_placed_length_df, 
        on=['PROJECT_NAME', 'WORK_ORDER_ID'], how='outer')
    
    merged_df = pd.merge(
        merged_df, 
        fiber_yesterday_length_df, 
        on=['PROJECT_NAME', 'WORK_ORDER_ID'], how='outer')
    
    merged_df = pd.merge(
        merged_df, 
        fiber_this_week_length_df, 
        on=['PROJECT_NAME', 'WORK_ORDER_ID'], how='outer')
    
    merged_df = pd.merge(
        merged_df, 
        fiber_last_week_length_df, 
        on=['PROJECT_NAME', 'WORK_ORDER_ID'], how='outer')
    
    merged_df = pd.merge(
        merged_df, 
        fiber_placed_length_df, 
        on=['PROJECT_NAME', 'WORK_ORDER_ID'], how='outer')
    
    print('calculating percents')
    merged_df['%_Conduit_Placed'] = round(merged_df['Conduit_Placed']/merged_df['Total_Length'] * 100)
    merged_df['%_Fiber_Placed'] = round(merged_df['Fiber_Placed']/merged_df['Total_Length'] * 100)

    """
    merged_df = pd.merge(
        merged_df, 
        total_demand_df, 
        on=['PROJECT_NAME', 'WORK_ORDER_ID'], how='outer')
    """
    
    # Write to excel workbook
    wb = pyxl.Workbook()
    ws = wb.active
    ws.title = 'Total'
    permit_sheet = wb.create_sheet(title="Cycle Times")
    activated_permit_sheet = wb.create_sheet(title="Activated Permits")

    i=1
    directory = r"C:\Users\cbailey\Desktop\const_updates_test"
    filename = "daily_update"
    save_path = os.path.join(directory, filename+'.xlsx')
    while os.path.exists(save_path):
        save_path = os.path.join(directory, filename+f'({i}).xlsx')
        i += 1
    wb.save(save_path)

    with pd.ExcelWriter(save_path, engine='openpyxl', mode='a', 
        if_sheet_exists='overlay') as writer:
        
        merged_df.to_excel(writer, sheet_name='Total', index=False, header=True)
        total_permit_df.to_excel(writer, sheet_name='Cycle Times', index=False, header=True)
        activated_permits.to_excel(writer, sheet_name='Activated Permits', index=False, header=True)

if __name__ == "__main__":
    main()